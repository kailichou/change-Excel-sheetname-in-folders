import os 
import io
import msoffcrypto
import openpyxl
import sys
from openpyxl import load_workbook
from pathlib import Path

cwd = os.getcwd()
root = os.path.abspath(os.path.join("..", os.pardir))
# path = cwd.rsplit('\\',1)[0]+'\input'
path = root + "BI_data\\input"
path1 = root + "BI_system\\bat"


with open(path1+'\clients.txt','r') as f:
    fold_names = f.readlines()

# arg = sys.argv[1]
# fold_lst = path+'\\'+ arg
fold_lst = []
for f in fold_names:
   fold_lst.append(path+'\\'+f.replace('\n','').encode("gbk").decode("utf-8"))


def getFiles(folder):
    dir_list = os.listdir(folder)
    filepaths = []
    for d in dir_list:
        if "FC.xlsx" in d:
            filepaths.append(folder+'\\'+d)
        elif "出货" in d:
            filepaths.append(folder+'\\'+d)
    return filepaths

def decrypt_wb(filepath):
    decrypted_workbook = io.BytesIO()
    with open(Path(filepath),'rb') as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password='20230630')
        office_file.decrypt(decrypted_workbook)
    return decrypted_workbook


def changeSheetName(filepath):
    if "出货" in filepath:
        decrypted_workbook=decrypt_wb(filepath)
        wb = load_workbook(filename=decrypted_workbook)
    else:
        wb = load_workbook(Path(filepath))
    #wb.security.lockStructure = False
    #wb.save(Path(filepath))
    #wb.close()
    sh_name = wb.sheetnames
    sh = wb [sh_name[0]]
    if sh.title == 'sheet1':
        pass
    else:
        sh.title='sheet1'
    wb.save(filepath)
    wb.close()


#files = getFiles(fold_lst[0])
for l in fold_lst:    
    paths = getFiles(l)
    #print(paths)
    for p in paths:
        changeSheetName(p)